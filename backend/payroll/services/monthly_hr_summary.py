"""Monthly HR summary Excel — salaries, payroll run figures, and leave balances."""

from __future__ import annotations

import io
from calendar import monthrange
from datetime import date
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from employees.models import Employee, Organization
from employees.org_scope import filter_employees_by_org
from leave_management.models import LeaveRequest, LeaveStatus, LeaveType
from leave_management.leave_rules import leave_days_in_year, quota_for, resolve_policy

from ..models import PayrollEmployeeResult, PayrollRun


def _leave_days_in_month(employee, leave_type: str, year: int, month: int) -> int:
    _, last = monthrange(year, month)
    month_start = date(year, month, 1)
    month_end = date(year, month, last)
    rows = LeaveRequest.objects.filter(
        employee=employee,
        status=LeaveStatus.APPROVED,
        leave_type=leave_type,
        start_date__lte=month_end,
        end_date__gte=month_start,
    ).only("start_date", "end_date")
    total = 0
    for row in rows:
        start = max(row.start_date, month_start)
        end = min(row.end_date, month_end)
        total += (end - start).days + 1
    return total


def _employee_name(employee: Employee) -> str:
    u = employee.user
    return f"{u.first_name} {u.last_name}".strip() or u.email


def _monthly_gross(employee: Employee, result: PayrollEmployeeResult | None) -> Decimal | None:
    if result and result.gross_monthly_full:
        return result.gross_monthly_full
    comp = getattr(employee, "compensation", None)
    if comp:
        if comp.monthly_gross:
            return comp.monthly_gross
        if comp.annual_ctc:
            return (comp.annual_ctc / Decimal("12")).quantize(Decimal("0.01"))
    return None


def _fmt_dec(value) -> str | float:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    return value


def _fmt_quota(quota) -> str | int:
    if quota is None:
        return "Unlimited"
    return int(quota)


def build_monthly_hr_summary_rows(org_id: int, year: int, month: int) -> list[dict]:
    emp_qs = Employee.objects.filter(
        user__is_active=True,
        user__onboarding_pending=False,
    ).select_related(
        "user",
        "compensation",
        "leave_policy_assignment",
        "leave_policy_assignment__policy",
    )
    emp_qs = filter_employees_by_org(emp_qs, org_id).order_by("employee_code")

    run = PayrollRun.objects.filter(
        organization_id=org_id,
        period_year=year,
        period_month=month,
    ).first()
    results_by_employee: dict[int, PayrollEmployeeResult] = {}
    if run:
        for res in PayrollEmployeeResult.objects.filter(run=run).select_related("employee"):
            results_by_employee[res.employee_id] = res

    rows: list[dict] = []
    for emp in emp_qs:
        policy, is_on_probation = resolve_policy(emp)
        result = results_by_employee.get(emp.id)
        gross = _monthly_gross(emp, result)

        leave_cols: dict[str, str | int] = {}
        if policy:
            for lt, prefix in (
                (LeaveType.CASUAL, "cl"),
                (LeaveType.PAID, "al"),
                (LeaveType.SICK, "sl"),
            ):
                quota = quota_for(policy, lt, is_on_probation, employee=emp, as_of=date(year, month, monthrange(year, month)[1]))
                used_ytd = leave_days_in_year(emp, lt, year)
                taken_month = _leave_days_in_month(emp, lt, year, month)
                remaining = "" if quota is None else max(int(quota) - used_ytd, 0)
                leave_cols[f"{prefix}_quota"] = _fmt_quota(quota)
                leave_cols[f"{prefix}_used_ytd"] = used_ytd
                leave_cols[f"{prefix}_balance"] = remaining if remaining != "" else "Unlimited"
                leave_cols[f"{prefix}_taken_month"] = taken_month
        else:
            for prefix in ("cl", "al", "sl"):
                leave_cols[f"{prefix}_quota"] = "—"
                leave_cols[f"{prefix}_used_ytd"] = "—"
                leave_cols[f"{prefix}_balance"] = "—"
                leave_cols[f"{prefix}_taken_month"] = "—"

        rows.append(
            {
                "employee_code": emp.employee_code,
                "name": _employee_name(emp),
                "department": emp.department or "",
                "designation": emp.designation or "",
                "monthly_gross": _fmt_dec(gross),
                "paid_days": _fmt_dec(result.paid_days) if result else "",
                "lop_days": _fmt_dec(result.lop_days) if result else "",
                "net_pay": _fmt_dec(result.net_pay) if result else "",
                **leave_cols,
            }
        )
    return rows


def export_monthly_hr_summary_xlsx(org_id: int, year: int, month: int) -> HttpResponse:
    org = Organization.objects.filter(pk=org_id).first()
    org_name = org.name if org else f"Organization {org_id}"
    period_label = date(year, month, 1).strftime("%B %Y")
    rows = build_monthly_hr_summary_rows(org_id, year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = "HR Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F766E")
    title_font = Font(bold=True, size=14)

    ws["A1"] = f"Monthly HR Summary — {org_name}"
    ws["A1"].font = title_font
    ws["A2"] = f"Period: {period_label}"
    ws["A3"] = f"Generated: {timezone.localtime().strftime('%d %b %Y, %I:%M %p')}"
    if not rows:
        ws["A5"] = "No active employees found for this organization."
    else:
        headers = [
            ("Employee Code", "employee_code"),
            ("Employee Name", "name"),
            ("Department", "department"),
            ("Designation", "designation"),
            ("Monthly Gross (₹)", "monthly_gross"),
            ("Paid Days", "paid_days"),
            ("LOP Days", "lop_days"),
            ("Net Pay (₹)", "net_pay"),
            ("CL Quota", "cl_quota"),
            ("CL Used (YTD)", "cl_used_ytd"),
            ("CL Balance", "cl_balance"),
            ("CL Taken (This Month)", "cl_taken_month"),
            ("Annual Quota", "al_quota"),
            ("Annual Used (YTD)", "al_used_ytd"),
            ("Annual Balance", "al_balance"),
            ("Annual Taken (This Month)", "al_taken_month"),
            ("Sick Quota", "sl_quota"),
            ("Sick Used (YTD)", "sl_used_ytd"),
            ("Sick Balance", "sl_balance"),
            ("Sick Taken (This Month)", "sl_taken_month"),
        ]
        start_row = 5
        for col_idx, (label, _) in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, row in enumerate(rows, start=start_row + 1):
            for col_idx, (_, key) in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            max_len = len(str(headers[col_idx - 1][0]))
            for row_idx in range(start_row + 1, start_row + 1 + len(rows)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 36)

        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"hr_monthly_summary_{year}_{month:02d}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
